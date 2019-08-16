import matplotlib
matplotlib.use('Agg')
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.ticker as tkr
import seaborn as sns
import os
from collections import OrderedDict
from openpyxl import load_workbook
import csv

color_air_water = ['palegreen', 'lightgreen', 'forestgreen', 'limegreen', 'aquamarine']
color_air = ['lightblue', 'dodgerblue']
color_water = ['palegreen', 'lightgreen', 'forestgreen', 'limegreen',
               'aquamarine', 'turquoise', 'lightseagreen', 'teal']
color_soil = ['bisque', 'turquoise', 'tan', 'darkgoldenrod', 'bisque', 'turquoise', 'tan', 'darkgoldenrod',
              'bisque', 'turquoise', 'tan', 'darkgoldenrod', 'bisque', 'turquoise', 'tan', 'darkgoldenrod']

class GenerateResult:

    def __init__(self):
        pass

    def store_output(self, chem_type, chem_name, region_name, release_scenario, release,
                     date_array, process_array, V_bulk_list, funC_df_list, funM_df_list, output_file_path, file_name):
        header = ['air', 'fw', 'fw_sed', 'sw', 'sw_sed', 'undeveloped_soil', 'deep_undeveloped_soil', 'urban_soil',
                           'deep_urban_soil', 'agricultural_soil', 'deep_agricultural_soil', 'biosolids_soil', 'deep_biosolids_soil'] # 13

        header_non_nano_sub = ['air', 'aerosol', 'fw', 'fw_sus_sed', 'fw_sed_water', 'fw_sed_solid', 'sw', 'sw_sus_sed',
                               'sw_sed_water', 'sw_sed_solid', 'undeveloped_soil_air', 'undeveloped_soil_water',
                               'undeveloped_soil_solid', 'deep_undeveloped_soil', 'urban_soil_air', 'urban_soil_water',
                               'urban_soil_solid', 'deep_urban_soil', 'agricultural_soil_air', 'agricultural_soil_water',
                               'agricultural_soil_solid', 'deep_agricultural_soil', 'biosolids_soil_air',
                               'biosolids_soil_water', 'biosolids_soil_solid', 'deep_biosolids_soil'] # 26

        # if funC_df_list only contains 2 items, it is NonionizableOrganic
        writer_C = pd.ExcelWriter(os.path.join(output_file_path, 'chem_conc_'+file_name+'.xlsx'))
        writer_M = pd.ExcelWriter(os.path.join(output_file_path, 'chem_mass_' + file_name + '.xlsx'))

        df_1_C = pd.DataFrame(funC_df_list[0], columns=header, index=date_array)
        df_1_M = pd.DataFrame(funM_df_list[0], columns=header, index=date_array)
        if chem_type != 'Nanomaterial':
            df_1_sub_C = pd.DataFrame(funC_df_list[1], columns=header_non_nano_sub, index=date_array)
            df_1_M = pd.DataFrame(funM_df_list[0], columns=header, index=date_array)
            df_1_sub_M = pd.DataFrame(funM_df_list[1], columns=header_non_nano_sub, index=date_array)

        self.generate_release_bulk(release, chem_name, region_name, release_scenario, os.path.join(output_file_path, 'release_bulk_'+file_name+'.png'))

        if chem_type == 'NonionizableOrganic' or chem_type == 'IonizableOrganic':
            txt = 'neutral'
            df_1_C.to_excel(writer_C, txt)
            df_1_sub_C.to_excel(writer_C, txt+'_sub')
            df_1_M.to_excel(writer_M, txt)
            df_1_sub_M.to_excel(writer_M, txt + '_sub')
            self.generate_plot_bulk(df_1_C, txt, chem_name, region_name, release_scenario, os.path.join(output_file_path, txt+'_bulk_'+file_name+'.png'))
            self.generate_mass_bulk(df_1_C, txt, V_bulk_list, chem_name, region_name, release_scenario, os.path.join(output_file_path, txt+'_bulk_mass_'+file_name+'.png'))
            self.generate_plot_sub(df_1_sub_C, txt, chem_name, chem_type, region_name, release_scenario, os.path.join(output_file_path, txt+'_sub_'+file_name+'.png'))

        else:
            # for metal and nanomaterial, no subcompartment
            txt = 'particulate'
            df_1_C.to_excel(writer_C, txt)
            df_1_M.to_excel(writer_M, txt)
            self.generate_plot_bulk(df_1_C, txt, chem_name, region_name, release_scenario, os.path.join(output_file_path, txt+'_bulk_'+file_name+'.png'))
            self.generate_mass_bulk(df_1_C, txt, V_bulk_list, chem_name, region_name, release_scenario,
                                    os.path.join(output_file_path, txt+'_bulk_mass_'+file_name+'.png'))

        if chem_type == 'IonizableOrganic':
            df_2_C = pd.DataFrame(funC_df_list[2], columns=header, index=date_array)
            df_2_sub_C = pd.DataFrame(funC_df_list[3], columns=header_non_nano_sub, index=date_array)
            df_2_M = pd.DataFrame(funM_df_list[2], columns=header, index=date_array)
            df_2_sub_M = pd.DataFrame(funM_df_list[3], columns=header_non_nano_sub, index=date_array)
            df_2_C.to_excel(writer_C, 'ionic')
            df_2_sub_C.to_excel(writer_C, 'ionic_sub')
            df_2_M.to_excel(writer_M, 'ionic')
            df_2_sub_M.to_excel(writer_M, 'ionic_sub')
            df_sum_bulk = df_1_C.add(df_2_C)
            df_sum_sub = df_1_sub_C.add(df_2_sub_C)
            txt = 'ionic'
            txt_sub = 'ionic_sub'
            self.generate_plot_bulk(df_2_C, txt, chem_name, region_name, release_scenario, os.path.join(output_file_path, txt+'_bulk_'+file_name+'.png'))
            self.generate_mass_bulk(df_2_C, txt, V_bulk_list, chem_name, region_name, release_scenario,
                                    os.path.join(output_file_path, txt+'_bulk_mass_'+file_name+'.png'))
            self.generate_plot_sub(df_2_sub_C, txt, chem_name, chem_type, region_name, release_scenario, os.path.join(output_file_path, txt_sub+'_'+file_name+'.png'))
            self.generate_plot_bulk(df_sum_bulk, '', chem_name, region_name, release_scenario, os.path.join(output_file_path, 'sum_bulk_'+file_name+'.png'))
            self.generate_mass_bulk(df_sum_bulk, '', V_bulk_list, chem_name, region_name, release_scenario,
                                    os.path.join(output_file_path, 'sum_bulk_mass_'+file_name+'.png'))

            self.generate_plot_sub(df_sum_sub, '', chem_name, chem_type, region_name, release_scenario, os.path.join(output_file_path, 'sum_sub_'+file_name+'.png'))
        elif chem_type == 'Metal' or chem_type == 'Nanomaterial':
            if chem_type == 'Metal':
                index1, index2 = 2, 4
                txt1, txt2 = 'colloidal', 'dissolved'
            else:
                index1, index2 = 1, 2
                txt1, txt2 = 'free nano', 'dissolved'

            df_2_C = pd.DataFrame(funC_df_list[index1], columns=header, index=date_array)
            df_2_M = pd.DataFrame(funM_df_list[index1], columns=header, index=date_array)
            df_3_C = pd.DataFrame(funC_df_list[index2], columns=header, index=date_array)
            df_3_M = pd.DataFrame(funM_df_list[index2], columns=header, index=date_array)
            df_2_C.to_excel(writer_C, txt1)
            df_3_C.to_excel(writer_C, txt2)
            df_2_M.to_excel(writer_M, txt1)
            df_3_M.to_excel(writer_M, txt2)
            df_sum1 = df_1_C.add(df_2_C)
            df_sum = df_sum1.add(df_3_C)

            self.generate_plot_bulk(df_2_C, txt1, chem_name, region_name, release_scenario, os.path.join(output_file_path, txt1+'_bulk_'+file_name+'.png'))
            self.generate_plot_bulk(df_3_C, txt2, chem_name, region_name, release_scenario, os.path.join(output_file_path, txt2+'_bulk_'+file_name+'.png'))
            self.generate_plot_bulk(df_sum, '', chem_name, region_name, release_scenario, os.path.join(output_file_path, 'sum_bulk_'+file_name+'.png'))

            self.generate_mass_bulk(df_2_C, txt1, V_bulk_list, chem_name, region_name, release_scenario,
                                    os.path.join(output_file_path, txt1+'_bulk_mass_'+file_name+'.png'))
            self.generate_mass_bulk(df_3_C, txt2, V_bulk_list, chem_name, region_name, release_scenario,
                                    os.path.join(output_file_path, txt2+'_bulk_mass_'+file_name+'.png'))
            self.generate_mass_bulk(df_sum, '', V_bulk_list, chem_name, region_name, release_scenario,
                                    os.path.join(output_file_path, 'sum_bulk_mass_'+file_name+'.png'))

        if chem_type == 'NonionizableOrganic':
            txt = 'organoFate'
            self.generate_heatmap_bulk(chem_type, [df_1_C], chem_name, region_name, release_scenario, V_bulk_list, os.path.join(output_file_path, txt+'_heatmap_'+file_name+'.png'))
        elif chem_type == 'IonizableOrganic':
            self.generate_heatmap_bulk(chem_type, [df_1_C, df_2_C], chem_name, region_name, release_scenario, V_bulk_list, os.path.join(output_file_path, 'ionOFate_heatmap_'+file_name+'.png'))
        else:
            if chem_type == 'Metal':
                txt = 'metal'
            else:
                txt = 'nano'
            self.generate_heatmap_bulk(chem_type, [df_1_C, df_2_C, df_3_C], chem_name, region_name, release_scenario, V_bulk_list, os.path.join(output_file_path, txt+'_heatmap_'+file_name+'.png'))

        writer_C.save()
        writer_M.save()
        print ('Saved the results to excel file.')

        # output process data
        self.store_process_output(chem_type, date_array, process_array, output_file_path, file_name)

        # insert unit into A1 cell in conc, mass and process file
        conc_file = os.path.join(output_file_path, 'chem_conc_' + file_name + '.xlsx')
        mass_file = os.path.join(output_file_path, 'chem_mass_' + file_name + '.xlsx')
        process_file = os.path.join(output_file_path, 'process_' + file_name + '.xlsx')

        conc_book = load_workbook(conc_file)
        conc_sheetnames = conc_book.sheetnames
        mass_book = load_workbook(mass_file)
        for sheetname in conc_sheetnames:
            conc_sheet = conc_book.get_sheet_by_name(sheetname)
            mass_sheet = mass_book.get_sheet_by_name(sheetname)
            conc_sheet['A1'] = 'g/L'
            mass_sheet['A1'] = 'kg'

        process_book = load_workbook(process_file)
        process_sheet = process_book.get_sheet_by_name('process')
        process_sheet['A1'] = 'kg/day'

        conc_book.save(os.path.join(output_file_path, 'chem_conc_' + file_name + '.xlsx'))
        mass_book.save(os.path.join(output_file_path, 'chem_mass_' + file_name + '.xlsx'))
        process_book.save(os.path.join(output_file_path, 'process_' + file_name + '.xlsx'))


    def store_process_output(self, chem_type, date_array, process_array, output_file_path, file_name):
        if chem_type == 'IonizableOrganic':
            header_list = ['adv_air_in', 'adv_air_out', 'adv_fw_in', 'adv_fw_out', 'adv_fwSed_in', 'adv_fwSed_out',
                           'adv_sw_in', 'adv_sw_out', 'adv_swSed_out', 'dep_dry_air', 'dep_dry_air_fw', 'dep_dry_air_sw',
                           'dep_dry_air_soil1', 'dep_dry_air_soil2', 'dep_dry_air_soil3', 'dep_dry_air_soil4',
                           'dep_wet_air', 'dep_wet_air_fw', 'dep_wet_air_sw', 'dep_wet_air_soil1', 'dep_wet_air_soil2',
                           'dep_wet_air_soil3', 'dep_wet_air_soil4', 'rain_dis_air', 'rain_dis_air_fw', 'rain_dis_air_sw',
                           'rain_dis_air_soil1', 'rain_dis_air_soil2', 'rain_dis_air_soil3', 'rain_dis_air_soil4',
                           'dep_fSS', 'dep_sSS', 'diff_air_fw', 'diff_air_sw', 'diff_fw_fSedW', 'diff_sw_sSedW',
                           'diff_air_soil1', 'diff_air_soil2', 'diff_air_soil3', 'diff_air_soil4', 'diff_fw_air',
                           'diff_sw_air', 'diff_fSedW_fw', 'diff_sSedW_sw', 'diff_soil1_air', 'diff_soil2_air',
                           'diff_soil3_air', 'diff_soil4_air', 'burial_fwSed', 'burial_swSed', 'resusp_fwSed',
                           'resusp_swSed', 'aero_resusp_sSS', 'runoff_soil1', 'runoff_soil2', 'runoff_soil3',
                           'runoff_soil4', 'erosion_soil1', 'erosion_soil2', 'erosion_soil3',
                           'erosion_soil4', 'wind_erosion_soil1', 'wind_erosion_soil2', 'wind_erosion_soil3',
                           'wind_erosion_soil4', 'infiltra_soil1', 'infiltra_soil2', 'infiltra_soil3', 'infiltra_soil4',
                           'leach_soil1', 'leach_soil2', 'leach_soil3', 'leach_soil4', 'deg_air', 'deg_fw', 'deg_fwSed',
                           'deg_sw', 'deg_swSed', 'deg_soil1', 'deg_deepS1', 'deg_soil2', 'deg_deepS2', 'deg_soil3',
                           'deg_deepS3', 'deg_soil4', 'deg_deepS4']
        elif chem_type == 'Metal':
            header_list = ['adv_air_in', 'adv_air_out', 'adv_fw_in', 'adv_fw_out', 'adv_fwSed_in', 'adv_fwSed_out',
                           'adv_sw_in', 'adv_sw_out', 'adv_swSed_out', 'dep_dry_air', 'dep_dry_air_fw', 'dep_dry_air_sw',
                           'dep_dry_air_soil1', 'dep_dry_air_soil2', 'dep_dry_air_soil3', 'dep_dry_air_soil4',
                           'dep_wet_air', 'dep_wet_air_fw', 'dep_wet_air_sw', 'dep_wet_air_soil1', 'dep_wet_air_soil2',
                           'dep_wet_air_soil3', 'dep_wet_air_soil4', 'dep_fSS', 'dep_sSS', 'diff_fw_fSedW', 'diff_sw_sSedW',
                           'diff_fSedW_fw', 'diff_sSedW_sw', 'burial_fwSed', 'burial_swSed', 'resusp_fwSed', 'resusp_swSed',
                           'aero_resusp_sSS', 'runoff_soil1', 'runoff_soil2', 'runoff_soil3', 'runoff_soil4', 'erosion_soil1',
                           'erosion_soil2', 'erosion_soil3', 'erosion_soil4', 'wind_erosion_soil1', 'wind_erosion_soil2',
                           'wind_erosion_soil3', 'wind_erosion_soil4', 'infiltra_soil1', 'infiltra_soil2', 'infiltra_soil3',
                           'infiltra_soil4', 'leach_soil1','leach_soil2', 'leach_soil3', 'leach_soil4']
        elif chem_type == 'NonionizableOrganic':
            header_list = ['adv_air_in', 'adv_air_out', 'adv_fw_in', 'adv_fw_out', 'adv_fwSed_in', 'adv_fwSed_out',
                           'adv_sw_in', 'adv_sw_out', 'adv_swSed_out', 'dep_dry_air', 'dep_dry_air_fw', 'dep_dry_air_sw',
                           'dep_dry_air_soil1', 'dep_dry_air_soil2', 'dep_dry_air_soil3', 'dep_dry_air_soil4',
                           'dep_wet_air', 'dep_wet_air_fw', 'dep_wet_air_sw', 'dep_wet_air_soil1', 'dep_wet_air_soil2',
                           'dep_wet_air_soil3', 'dep_wet_air_soil4', 'rain_dis_air', 'rain_dis_air_fw', 'rain_dis_air_sw',
                           'rain_dis_air_soil1', 'rain_dis_air_soil2', 'rain_dis_air_soil3', 'rain_dis_air_soil4',
                           'dep_fSS', 'dep_sSS', 'diff_air_fw', 'diff_air_sw', 'diff_fw_fSedW', 'diff_sw_sSedW',
                           'diff_air_soil1', 'diff_air_soil2', 'diff_air_soil3', 'diff_air_soil4', 'diff_fw_air',
                           'diff_sw_air', 'diff_fSedW_fw', 'diff_sSedW_sw', 'diff_soil1_air', 'diff_soil2_air',
                           'diff_soil3_air', 'diff_soil4_air', 'burial_fwSed', 'burial_swSed', 'resusp_fwSed',
                           'resusp_swSed', 'runoff_soil1', 'runoff_soil2', 'runoff_soil3',
                           'runoff_soil4', 'erosion_soil1', 'erosion_soil2', 'erosion_soil3',
                           'erosion_soil4', 'wind_erosion_soil1', 'wind_erosion_soil2', 'wind_erosion_soil3',
                           'wind_erosion_soil4', 'infiltra_soil1', 'infiltra_soil2', 'infiltra_soil3', 'infiltra_soil4',
                           'leach_soil1', 'leach_soil2', 'leach_soil3', 'leach_soil4', 'deg_air', 'deg_fw', 'deg_fwSed',
                           'deg_sw', 'deg_swSed', 'deg_soil1', 'deg_deepS1', 'deg_soil2', 'deg_deepS2', 'deg_soil3',
                           'deg_deepS3', 'deg_soil4', 'deg_deepS4']
        else:
            header_list = ['adv_air_in', 'adv_fw_out', 'adv_fwSed_out', 'adv_sw_out', 'adv_swSed_out', 'adv_fw_dissolved',
                           'adv_fwSed_dissolved', 'adv_sw_dissolved', 'adv_swSed_dissolved', 'dep_dry_air', 'dep_dry_air_fw',
                           'dep_dry_air_sw', 'dep_dry_air_soil1', 'dep_dry_air_soil2', 'dep_dry_air_soil3', 'dep_dry_air_soil4',
                           'dep_wet_air', 'dep_wet_air_fw', 'dep_wet_air_sw', 'dep_wet_air_soil1', 'dep_wet_air_soil2',
                           'dep_wet_air_soil3', 'dep_wet_air_soil4', 'dep_fSS', 'dep_sSS', 'heteroagg_air', 'heteroagg_fw',
                           'heteroagg_sw', 'dissolution_fw', 'dissolution_fwSed', 'dissolution_sw', 'dissolution_swSed',
                           'dissolution_soil1', 'dissolution_soil2', 'dissolution_soil3', 'dissolution_soil4',
                           'partition_soil2soilw1', 'partition_soil2soilw2', 'partition_soil2soilw3', 'partition_soil2soilw4',
                           'partition_soilw2soil1', 'partition_soilw2soil2', 'partition_soilw2soil3', 'partition_soilw2soil4',
                           'runoff_soil1', 'runoff_soil2', 'runoff_soil3', 'runoff_soil4', 'runoff_soil1_dissolved',
                           'runoff_soil2_dissolved', 'runoff_soil3_dissolved', 'runoff_soil4_dissolved', 'erosion_soil1',
                           'erosion_soil2', 'erosion_soil3', 'erosion_soil4', 'wind_erosion_soil1', 'wind_erosion_soil2',
                           'wind_erosion_soil3', 'wind_erosion_soil4', 'infiltra_soil1', 'infiltra_soil2', 'infiltra_soil3',
                           'infiltra_soil4', 'leach_soil1', 'leach_soil2', 'leach_soil3', 'leach_soil4',
                           'burial_fwSed', 'burial_swSed', 'resusp_fwSed', 'resusp_swSed', 'aero_resusp_sSS']

        writer = pd.ExcelWriter(os.path.join(output_file_path, 'process_' + file_name + '.xlsx'))
        df = pd.DataFrame(process_array, columns = header_list, index = date_array)
        df.to_excel(writer, 'process')
        writer.save()


    def generate_plot_bulk(self, df, txt, chem_name, region_name, release_scenario, output_figure_path):
        df = df.reset_index(drop=True)
        sns.set(style='white')
        sim_days = df.shape[0]
        X = np.linspace(0, sim_days, num=sim_days)
        fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(8, 5), sharey=True)

        line_labels = ['Air', 'FW', 'FW Sed', 'SW', 'SW Sed', 'Natural Soil',
                       'Urban Soil', 'Ag. Soil', 'Bio. Soil']
        sns.despine()

        l1 = axes[0].plot(X, df['air'].tolist(), color='lightsteelblue')[0]
        l2 = axes[0].plot(X, df['fw'].tolist(), color='deepskyblue')[0]
        l3 = axes[0].plot(X, df['fw_sed'].tolist(), color='deepskyblue', linestyle=':')[0]
        l4 = axes[0].plot(X, df['sw'].tolist(), color='royalblue')[0]
        l5 = axes[0].plot(X, df['sw_sed'].tolist(), color='royalblue', linestyle=':')[0]

        l6 = axes[1].plot(X, df['undeveloped_soil'].tolist(), color='goldenrod')[0]
        l7 = axes[1].plot(X, df['urban_soil'].tolist(), color='darkgrey')[0]
        l8 = axes[1].plot(X, df['agricultural_soil'].tolist(), color='limegreen')[0]
        l9 = axes[1].plot(X, df['biosolids_soil'].tolist(), color='m')[0]

        axes[0].set_yscale('log')
        axes[1].set_yscale('log')

        axes[0].set_xlim(xmin=0)
        axes[1].set_xlim(xmin=0)

        axes[0].set_ylabel('Concentration (g/L)', fontsize=14)

        axes[0].set_title("Air and Water")
        axes[1].set_title('Soil')

        fig.suptitle(txt.capitalize() + ' ' + chem_name + ' in ' + region_name + ' using ' + release_scenario + ' Release \n'
                     'Concentration in Major (Bulk) Compartments')

        fig.text(0.5, 0.01, 'Simulation Time (day)', ha='center', fontsize=14)

        fig.legend([l1, l2, l3, l4, l5, l6, l7, l8, l9],
                   labels=line_labels,
                   loc='center right', frameon=False)

        plt.subplots_adjust(right=0.75, top=0.8)
        plt.savefig(output_figure_path)
        plt.close('all')


    def generate_plot_sub(self, df, txt, chem_name, chem_type, region_name, release_scenario, output_figure_path):
        df = df.reset_index(drop=True)
        sns.set(style='white')
        sim_days = df.shape[0]
        fig, axes = plt.subplots(nrows=1, ncols=3, figsize=(10, 7), sharey=True)
        sns.despine()
        X = np.linspace(0, sim_days, num=sim_days)

        if chem_type != 'Nanomaterial':
            line_label_non_nano = ['Air', 'Aero', 'FW', 'FW Susp. Sed', 'FW Sed Water', 'FW Sed Solid',
                                   'SW', 'SW Susp. Sed', 'SW Sed Water', 'SW Sed Solid',
                                   'Natural Soil Air', 'Natural Soil Water', 'Natural Soil Solid',
                                   'Natural Deep Soil', 'Urban Soil Air', 'Urban Soil Water', 'Urban Soil Solid', 'Urban Deep Soil',
                                   'Ag. Soil Air', 'Ag. Soil Water', 'Ag. Soil Solid', 'Ag. Deep Soil',
                                   'Ag. Deep Soil', 'Bio. Soil Air', 'Bio. Soil Water', 'Bio. Soil Solid', 'Bio. Deep Soil']


            l1 = axes[0].plot(X, df['air'].tolist(), color='lightsteelblue')[0]
            l2 = axes[0].plot(X, df['aerosol'].tolist(), color='steelblue')[0]
            l3 = axes[1].plot(X, df['fw'].tolist(), color='skyblue')[0]
            l4 = axes[1].plot(X, df['fw_sus_sed'].tolist(), color='skyblue', linestyle=':')[0]
            l5 = axes[1].plot(X, df['fw_sed_water'].tolist(), color='deepskyblue')[0]
            l6 = axes[1].plot(X, df['fw_sed_solid'].tolist(), color='deepskyblue', linestyle=':')[0]
            l7 = axes[1].plot(X, df['sw'].tolist(), color='blue')[0]
            l8 = axes[1].plot(X, df['sw_sus_sed'].tolist(), color='blue', linestyle=':')[0]
            l9 = axes[1].plot(X, df['sw_sed_water'].tolist(), color='slategray')[0]
            l10 = axes[1].plot(X, df['sw_sed_solid'].tolist(), color='slategray', linestyle=':')[0]

            l11 = axes[2].plot(X, df['undeveloped_soil_air'].tolist(), color='wheat')[0]
            l12 = axes[2].plot(X, df['undeveloped_soil_water'].tolist(), color='gold')[0]
            l13 = axes[2].plot(X, df['undeveloped_soil_solid'].tolist(), color='goldenrod')[0]
            l14 = axes[2].plot(X, df['deep_agricultural_soil'].tolist(), color='goldenrod', linestyle=':')[0]

            l15 = axes[2].plot(X, df['urban_soil_air'].tolist(), color='lightgray')[0]
            l16 = axes[2].plot(X, df['urban_soil_water'].tolist(), color='darkgrey')[0]
            l17 = axes[2].plot(X, df['urban_soil_solid'].tolist(), color='dimgray')[0]
            l18 = axes[2].plot(X, df['deep_urban_soil'].tolist(), color='dimgray', linestyle=':')[0]

            l19 = axes[2].plot(X, df['agricultural_soil_air'].tolist(), color='lightgreen')[0]
            l20 = axes[2].plot(X, df['agricultural_soil_water'].tolist(), color='limegreen')[0]
            l21 = axes[2].plot(X, df['agricultural_soil_solid'].tolist(), color='darkgreen')[0]
            l22 = axes[2].plot(X, df['deep_agricultural_soil'].tolist(), color='darkgreen', linestyle=':')[0]

            l23 = axes[2].plot(X, df['biosolids_soil_air'].tolist(), color='plum')[0]
            l24 = axes[2].plot(X, df['biosolids_soil_water'].tolist(), color='m')[0]
            l25 = axes[2].plot(X, df['biosolids_soil_solid'].tolist(), color='purple')[0]
            l26 = axes[2].plot(X, df['deep_biosolids_soil'].tolist(), color='purple', linestyle=':')[0]

            fig.legend([l1, l2, l3, l4, l5, l6, l7, l8, l9, l10, l11, l12, l13, l14, l15, l16, l17, l18,
                        l19, l20, l21, l22, l23, l24, l25, l26],
                       labels=line_label_non_nano,
                       loc='center right', frameon=False)

        else:
            line_label_nano = ['Air', 'Aero', 'FW', 'FW Susp. Sed', 'FW Dis.', 'FW Sed Solid', 'FW Sed Dis.',
                               'SW', 'SW Susp. Sed', 'SW Dis.', 'SW Sed Solid', 'SW Sed Dis.',
                               'Natural Soil Water', 'Natural Soil Solid', 'Natural Deep Soil', 'Natural Soil Dis.',
                               'Urban Soil Water', 'Urban Soil Solid', 'Urban Deep Soil', 'Urban Soil Dis.',
                               'Ag. Soil Water', 'Ag. Soil Solid', 'Ag. Deep Soil', 'Ag. Soil Dis.',
                               'Bio. Soil Water', 'Bio. Soil Solid', 'Bio. Deep Soil', 'Bio. Soil Dis.']
            l1 = axes[0].plot(X, df['air'].tolist(), color='lightsteelblue')[0]
            l2 = axes[0].plot(X, df['aerosol'].tolist(), color='steelblue')[0]
            l3 = axes[1].plot(X, df['fw'].tolist(), color='skyblue')[0]
            l4 = axes[1].plot(X, df['fw_sus_sed'].tolist(), color='deepskyblue')[0]
            l5 = axes[1].plot(X, df['fw_dis'].tolist(), color='skyblue', linestyle=':')[0]
            l6 = axes[1].plot(X, df['fw_sed'].tolist(), color='darkturquoise')[0]
            l7 = axes[1].plot(X, df['fwSed_dis'].tolist(), color='darkcyan', linestyle=':')[0]
            l8 = axes[1].plot(X, df['sw'].tolist(), color='cornflowerblue')[0]
            l9 = axes[1].plot(X, df['sw_sus_sed'].tolist(), color='blue')[0]
            l10 = axes[1].plot(X, df['sw_dis'].tolist(), color='royalblue', linestyle=':')[0]
            l11 = axes[1].plot(X, df['sw_sed'].tolist(), color='slategray')[0]
            l12 = axes[1].plot(X, df['swSed_dis'].tolist(), color='slategray', linestyle=':')[0]

            l13 = axes[2].plot(X, df['undeveloped_soil_water'].tolist(), color='wheat')[0]
            l14 = axes[2].plot(X, df['undeveloped_soil_solid'].tolist(), color='gold')[0]
            l15 = axes[2].plot(X, df['deep_undeveloped_soil'].tolist(), color='goldenrod')[0]
            l16 = axes[2].plot(X, df['undevSoilW_dis'].tolist(), color='goldenrod', linestyle=':')[0]

            l17 = axes[2].plot(X, df['urban_soil_water'].tolist(), color='lightgray')[0]
            l18 = axes[2].plot(X, df['urban_soil_solid'].tolist(), color='darkgray')[0]
            l19 = axes[2].plot(X, df['deep_urban_soil'].tolist(), color='dimgray')[0]
            l20 = axes[2].plot(X, df['urbanSoilW_dis'].tolist(), color='dimgray', linestyle=':')[0]

            l21 = axes[2].plot(X, df['agricultural_soil_water'].tolist(), color='lightgreen')[0]
            l22 = axes[2].plot(X, df['agricultural_soil_solid'].tolist(), color='limegreen')[0]
            l23 = axes[2].plot(X, df['deep_agricultural_soil'].tolist(), color='darkgreen')[0]
            l24 = axes[2].plot(X, df['agSoilW_dis'].tolist(), color='darkgreen', linestyle=':')[0]

            l25 = axes[2].plot(X, df['biosolids_soil_water'].tolist(), color='plum')[0]
            l26 = axes[2].plot(X, df['biosolids_soil_solid'].tolist(), color='m')[0]
            l27 = axes[2].plot(X, df['deep_biosolids_soil'].tolist(), color='purple')[0]
            l28 = axes[2].plot(X, df['bioSoilW_dis'].tolist(), color='purple', linestyle=':')[0]

            fig.legend([l1, l2, l3, l4, l5, l6, l7, l8, l9, l10, l11, l12, l13, l14, l15, l16, l17, l18, l19, l20,
                        l21, l22, l23, l24, l25, l26, l27, l28],
                       labels=line_label_nano,
                       loc='center right', frameon=False)

        axes[0].set_yscale('log')
        axes[1].set_yscale('log')
        axes[2].set_yscale('log')

        axes[0].set_ylabel('Concentration (g/L)', fontsize=14)

        axes[0].set_xlim(xmin=0)
        axes[1].set_xlim(xmin=0)
        axes[2].set_xlim(xmin=0)

        axes[0].set_title("Air")
        axes[1].set_title('Water')
        axes[2].set_title('Soil')

        fig.suptitle(txt.capitalize() + ' ' + chem_name + ' in ' + region_name + ' using ' + release_scenario + ' Release \n'
                                                                                       'Concentration in Subcompartments')

        fig.text(0.5, 0.01, 'Simulation Time (day)', ha='center', fontsize=14)
        plt.subplots_adjust(right=0.75, top=0.8)
        # plt.tight_layout()
        plt.savefig(output_figure_path)
        plt.close('all')



    def generate_heatmap_bulk(self, chem_type, df, chem_name, region_name, release_scenario,
                              V_bulk_list, output_figure_path):
        data = OrderedDict()
        data['conc'] = OrderedDict()
        data['massFr'] = OrderedDict()

        total_mass = 0

        for i in range(0,len(df)):
            if i == 0:
                txt = 'first'
            elif i == 1:
                txt = 'second'
            else:
                txt = 'third'

            data['conc'][txt] = OrderedDict()

            data['conc'][txt]['Air'] = df[i]['air'].mean()
            data['conc'][txt]['FW'] = df[i]['fw'].mean()
            data['conc'][txt]['FW Sed'] = df[i]['fw_sed'].mean()
            data['conc'][txt]['SW'] = df[i]['sw'].mean()
            data['conc'][txt]['SW Sed'] = df[i]['sw_sed'].mean()
            data['conc'][txt]['Natural Soil'] = df[i]['undeveloped_soil'].mean()
            data['conc'][txt]['Urban Soil'] = df[i]['urban_soil'].mean()
            data['conc'][txt]['Ag. Soil'] = df[i]['agricultural_soil'].mean()
            data['conc'][txt]['Bio. Soil'] = df[i]['biosolids_soil'].mean()

            total_mass = total_mass + df[i]['air'][-1] * V_bulk_list[0] + df[i]['fw'][-1] * V_bulk_list[1] + df[i]['fw_sed'][-1] * V_bulk_list[2] \
                         + df[i]['sw'][-1] * V_bulk_list[3] + df[i]['sw_sed'][-1] * V_bulk_list[4] + \
                         df[i]['undeveloped_soil'][-1] * V_bulk_list[5] + df[i]['urban_soil'][-1] * V_bulk_list[6] + \
                         df[i]['agricultural_soil'][-1] * V_bulk_list[7] + df[i]['biosolids_soil'][-1] * V_bulk_list[8]

        for i in range(0, len(df)):
            if i == 0:
                txt = 'first'
            elif i == 1:
                txt = 'second'
            else:
                txt = 'third'

            data['massFr'][txt] = OrderedDict()
            data['massFr'][txt]['Air'] = (df[i]['air'][-1] * V_bulk_list[0]/total_mass) * 100
            data['massFr'][txt]['FW'] = (df[i]['fw'][-1] * V_bulk_list[1]/total_mass) * 100
            data['massFr'][txt]['FW Sed'] = (df[i]['fw_sed'][-1] * V_bulk_list[2]/total_mass) * 100
            data['massFr'][txt]['SW'] = (df[i]['sw'][-1] * V_bulk_list[3]/total_mass) * 100
            data['massFr'][txt]['SW Sed'] = (df[i]['sw_sed'][-1] * V_bulk_list[4]/total_mass) * 100
            data['massFr'][txt]['Natural Soil'] = (df[i]['undeveloped_soil'][-1] * V_bulk_list[5]/total_mass) * 100
            data['massFr'][txt]['Urban Soil'] = (df[i]['urban_soil'][-1] * V_bulk_list[6]/total_mass) * 100
            data['massFr'][txt]['Ag. Soil'] = (df[i]['agricultural_soil'][-1] * V_bulk_list[7]/total_mass) * 100
            data['massFr'][txt]['Bio. Soil'] = (df[i]['biosolids_soil'][-1] * V_bulk_list[8]/total_mass) * 100

        df_conc = pd.DataFrame.from_dict(data['conc'])
        df_massFr = pd.DataFrame.from_dict(data['massFr'])

        df_conc = df_conc.reindex(['Air', 'FW', 'FW Sed', 'SW', 'SW Sed', 'Natural Soil', 'Urban Soil', 'Ag. Soil', 'Bio. Soil'])
        df_massFr = df_massFr.reindex(['Air', 'FW', 'FW Sed', 'SW', 'SW Sed', 'Natural Soil', 'Urban Soil', 'Ag. Soil', 'Bio. Soil'])

        color_heatmap = ['#ffffcc', '#feb24c', '#fc4e2a', '#b10026']
        cmap = mcolors.LinearSegmentedColormap.from_list("color", color_heatmap)

        if len(df) == 1:
            fig_width = 8
            fig_height = 5
        elif len(df) == 2:
            fig_width = 8
            fig_height = 6
        else:
            fig_width = 11
            fig_height = 6

        fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(fig_width, fig_height))

        if chem_type == 'NonionizableOrganic':
            xlabels = ['Neutral']
        elif chem_type == 'IonizableOrganic':
            xlabels = ['Neutral', 'Ionic']
        elif chem_type == 'Nanomaterial':
            xlabels = ['Particulate', 'Free Nano', 'Dissolved']
        else:
            xlabels = ['Particulate', 'Colloidal', 'Dissolved']

        sns.set(font_scale=1.1)
        sns.heatmap(df_conc, ax=axes[0], cmap=cmap,
                    cbar_kws={'label': 'Concentration (g/L)', 'format': '%.0e'},
                    linecolor='white', linewidths=0.5, annot=True, fmt='.1e', annot_kws={"size": 11})

        sns.heatmap(df_massFr, ax=axes[1], cmap=cmap,
                    cbar_kws={'label': 'Mass. Frac (%)'},
                    linecolor='white', linewidths=0.5, annot=True, fmt='.1e', annot_kws={"size": 11})

        axes[0].set_title("Mean Conc.")
        axes[1].set_title("Mass Frac.")

        axes[0].set_xticklabels(xlabels, fontsize=12)
        axes[1].set_xticklabels(xlabels, fontsize=12)

        fig.suptitle(chem_name + ' in ' + region_name + ' using ' + release_scenario + ' Release \n'
                     'Overall mean concentration and last day mass fraction in major (bulk) compartments \n', fontsize=13)

        plt.subplots_adjust(wspace=0.8, top=0.8)
        # plt.tight_layout()
        plt.savefig(output_figure_path)
        plt.close('all')


    def generate_release_bulk(self, release, chem_name, region_name, release_scenario, output_figure_path):
        air = release['air']
        fw = release['fw']
        fSS = release['fSS']
        fwSed = release['fwSed']
        sw = release['sw']
        sSS = release['sSS']
        swSed = release['swSed']
        soil1 = release['soil1']
        dsoil1 = release['dsoil1']
        soil2 = release['soil2']
        dsoil2 = release['dsoil2']
        soil3 = release['soil3']
        dsoil3 = release['dsoil3']
        soil4 = release['soil4']
        dsoil4 = release['dsoil4']

        compart_list = [air, fw, fSS, fwSed, sw, sSS, swSed, soil1, dsoil1, soil2, dsoil2, soil3, dsoil3, soil4, dsoil4]
        compart2_list = compart_list[:]

        for i in range(0, len(compart_list)):
            for j in range(0, len(compart_list[i])):
                if j == 0:
                    compart2_list[i][j] = compart_list[i][j]
                else:
                    compart2_list[i][j] = compart2_list[i][j-1] + compart_list[i][j]

        sns.set(style='white')
        sim_days = len(air)
        X = np.linspace(0, sim_days, num=sim_days)
        fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(8, 5), sharey=True)

        line_labels = ['Air', 'FW', 'FW Sus Sed', 'FW Sed', 'SW', 'SW Sus Sed', 'SW Sed', 'Natual Soil', 'Natural Deep Soil',
                       'Urban Soil', 'Urban Deep Soil', 'Ag. Soil', 'Ag. Deep Soil', 'Bio. Soil', 'Bio. Deep Soil']
        sns.despine()

        l1 = axes[0].plot(X, compart2_list[0], color='lightsteelblue')[0]
        l2 = axes[0].plot(X, compart2_list[1], color='skyblue')[0]
        l3 = axes[0].plot(X, compart2_list[2], color='deepskyblue')[0]
        l4 = axes[0].plot(X, compart2_list[3], color='deepskyblue', linestyle=':')[0]
        l5 = axes[0].plot(X, compart2_list[4], color='royalblue')[0]
        l6 = axes[0].plot(X, compart2_list[5], color='darkslateblue')[0]
        l7 = axes[0].plot(X, compart2_list[6], color='darkslateblue', linestyle=':')[0]

        l8 = axes[1].plot(X, compart2_list[7], color='goldenrod')[0]
        l9 = axes[1].plot(X, compart2_list[8], color='goldenrod', linestyle=':')[0]
        l10 = axes[1].plot(X, compart2_list[9], color='darkgray')[0]
        l11 = axes[1].plot(X, compart2_list[10], color='darkgray', linestyle=':')[0]
        l12 = axes[1].plot(X, compart2_list[11], color='limegreen')[0]
        l13 = axes[1].plot(X, compart2_list[12], color='limegreen', linestyle=':')[0]
        l14 = axes[1].plot(X, compart2_list[13], color='purple')[0]
        l15 = axes[1].plot(X, compart2_list[14], color='purple', linestyle=':')[0]

        # axes[0].set_yscale('log')
        # axes[1].set_yscale('log')

        plt.ticklabel_format(style='sci', axis='y', scilimits=(0, 0))

        axes[0].set_xlim(xmin=0)
        axes[1].set_xlim(xmin=0)

        axes[0].set_ylim(ymin=0)
        axes[0].set_ylim(ymin=0)

        axes[0].set_ylabel('Accumulated Release Amount (kg)', fontsize=14)

        axes[0].set_title("Air and Water")
        axes[1].set_title('Soil')

        fig.suptitle(chem_name + ' in ' + region_name + ' using ' + release_scenario + ' Release \n'
                     'Cumulative Mass Released to Major (Bulk) Compartments ')

        fig.text(0.5, 0.01, 'Simulation Time (day)', ha='center', fontsize=14)

        fig.legend([l1, l2, l3, l4, l5, l6, l7, l8, l9, l10, l11, l12, l13, l14, l15],
                   labels=line_labels,
                   loc='center right', frameon=False)

        plt.subplots_adjust(right=0.75, top=0.8)
        plt.savefig(output_figure_path)
        plt.close('all')


    def generate_mass_bulk(self, df, txt, V_bulk_list, chem_name, region_name, release_scenario, output_figure_path):
        df = df.reset_index(drop=True)
        sns.set(style='white')
        sim_days = df.shape[0]
        X = np.linspace(0, sim_days, num=sim_days)
        fig, axes = plt.subplots(nrows=1, ncols=2, figsize=(8, 5), sharey=True)

        line_labels = ['Air', 'FW', 'FW Sed', 'SW', 'SW Sed', 'Natual Soil',
                       'Urban Soil', 'Ag. Soil', 'Bio. Soil']
        sns.despine()

        l1 = axes[0].plot(X, (df['air']*V_bulk_list[0]).tolist(), color='lightsteelblue')[0]
        l2 = axes[0].plot(X, (df['fw']*V_bulk_list[1]).tolist(), color='deepskyblue')[0]
        l3 = axes[0].plot(X, (df['fw_sed']*V_bulk_list[2]).tolist(), color='deepskyblue', linestyle=':')[0]
        l4 = axes[0].plot(X, (df['sw']*V_bulk_list[3]).tolist(), color='royalblue')[0]
        l5 = axes[0].plot(X, (df['sw_sed']*V_bulk_list[4]).tolist(), color='royalblue', linestyle=':')[0]

        l6 = axes[1].plot(X, (df['undeveloped_soil']*V_bulk_list[5]).tolist(), color='goldenrod')[0]
        l7 = axes[1].plot(X, (df['urban_soil']*V_bulk_list[6]).tolist(), color='darkgrey')[0]
        l8 = axes[1].plot(X, (df['agricultural_soil']*V_bulk_list[7]).tolist(), color='limegreen')[0]
        l9 = axes[1].plot(X, (df['biosolids_soil']*V_bulk_list[8]).tolist(), color='m')[0]

        axes[0].set_yscale('log')
        axes[1].set_yscale('log')

        axes[0].set_xlim(xmin=0)
        axes[1].set_xlim(xmin=0)

        axes[0].set_ylabel('Mass (kg)', fontsize=14)

        axes[0].set_title("Air and Water")
        axes[1].set_title('Soil')

        fig.suptitle(txt.capitalize() + ' ' + chem_name + ' in ' + region_name + ' using ' + release_scenario + ' Release \n'
                     'Mass in Major (Bulk) Compartments')

        fig.text(0.5, 0.01, 'Simulation Time (day)', ha='center', fontsize=14)

        fig.legend([l1, l2, l3, l4, l5, l6, l7, l8, l9],
                   labels=line_labels,
                   loc='center right', frameon=False)

        plt.subplots_adjust(right=0.75, top=0.8)
        plt.savefig(output_figure_path)
        plt.close('all')
